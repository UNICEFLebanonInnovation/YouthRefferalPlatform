
from referral_platform.taskapp.celery import app

import json
import httplib
import datetime
from time import mktime
from django.core.serializers.json import DjangoJSONEncoder

from openpyxl import load_workbook
from referral_platform.youth.utils import generate_id
from .models import AssessmentSubmission, Assessment, Registration
from referral_platform.initiatives.models import YouthLedInitiative


@app.task
def import_registrations(filename, base_url, token, protocol='HTTPS'):

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet2']
    data = {}
    header = []
    index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            header.append((index, cell.value))
            index += 1

    for row in ws.rows:
        try:
            if row[0].value == 'start':
                continue
            data = {}
            # not for use 1, 2, 3, 4
            data['youth_nationality'] = int(row[5].value) if row[5].value else None
            data['governorate'] = int(row[6].value) if row[6].value else None
            data['partner_organization'] = int(row[7].value) if row[7].value else None
            # training type 8
            # center type 9
            if row[10].value or not row[10].value == 'na':
                data['center_id'] = row[10].value

            # concent paper 11
            # If_you_answered_Othe_the_name_of_the_NGO 12

            data['youth_first_name'] = row[13].value
            data['youth_last_name'] = row[14].value
            data['youth_father_name'] = row[15].value
            # nationality 1 instead of 16
            if row[16].value:
                data['youth_nationality'] = int(row[16].value)

            data['id_number'] = row[17].value  # UNHCR ID
            if not row[17].value:
                data['id_number'] = row[18].value  # Jordanian ID
            data['youth_bayanati_id'] = row[19].value  # Bayanati_ID
            # training date 20
            # training end date 21
            data['youth_sex'] = row[22].value
            if not row[23].value or row[23].value == 'na':
                data['youth_marital_status'] = 'single'
            else:
                data['youth_marital_status'] = row[23].value

            data['youth_birthday_year'] = ''
            data['youth_birthday_month'] = ''
            data['youth_birthday_day'] = ''
            if row[24].value:
                birthday = row[24].value.split('-')
                data['youth_birthday_year'] = int(birthday[0])
                data['youth_birthday_month'] = int(birthday[1])
                data['youth_birthday_day'] = int(birthday[2])

            result = post_data(protocol=protocol, url=base_url, apifunc='/api/registration/', token=token, data=data)
            registry = json.loads(result)

            submit_assessment(header, row, 1, registry, base_url, token, protocol)

        except Exception as ex:
            print("---------------")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            pass
##############################################################################
@app.task
def import_assessment_submission(filename):

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet1']
    new_data = {}
    header = []
    index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            header.append((index, cell.value))
            index += 1

    for row in ws.rows:
        try:
            if row[0].value == 'ID':
                continue
            new_data = {}
            # Rendering the assessment from excel

            for key, value in header:
                new_data[value] = row[key].value
                instance = AssessmentSubmission.objects.get(youth__number=row[0].value, assessment__slug=row[1].value)
                instance.new_data = new_data
                instance.updated = '1'
                print(instance)
                instance.save()

        except Exception as ex:
            print("---------------")
            print("error: ", ex.message)
            print(json.dumps(new_data, cls=DjangoJSONEncoder))
            print("---------------")
            pass


#################################################################################

@app.task
def import_initiatives(filename):

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet1']
    header = []
    index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            header.append((index, cell.value))
            index += 1

    for row in ws.rows:
        try:
            if row[0].value == 'ID':
                continue
            data = {}
            # not for use 1, 2, 3, 4
            obj = YouthLedInitiative(
                                     title=row[1].value,
                                     location=row[2].value,
                                     partner_organization=row[3].value,
                                     Participants=row[4].value,
                                     duration=row[5].value,
                                     type=row[6].value)
            print(obj)
            obj.save()


        except Exception as ex:
            print("---------------")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            pass
##############################################################################
@app.task
def update_registrations(filename, base_url, token, protocol='HTTPS'):
    from referral_platform.registrations.models import Registration

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet2']
    data = {}

    for row in ws.rows:
        try:
            if row[0].value == 'ID':
                continue
            data = {}

            data['youth_birthday_year'] = ''
            data['youth_birthday_month'] = ''
            data['youth_birthday_day'] = ''
            birthday_year = ''
            if row[24].value:
                birthday = row[24].value.split('-')
                birthday_year = int(birthday[0])
                data['youth_birthday_year'] = int(birthday[0])
                data['youth_birthday_month'] = int(birthday[1])
                data['youth_birthday_day'] = int(birthday[2])

            matched_registred_youth = Registration.objects.get(
                partner_organization_id=int(row[7].value),
                youth__first_name=row[13].value,
                youth__father_name=row[15].value,
                youth__last_name=row[14].value,
                youth__sex=row[22].value,
                youth__birthday_year=birthday_year
            )

            data['trainer'] = 'na'
            nationality = int(row[5].value)
            if nationality == 8:
                nationality = 2

            data['youth_nationality'] = nationality if nationality else None
            data['youth_first_name'] = row[13].value
            data['youth_last_name'] = row[14].value
            data['youth_father_name'] = row[15].value
            data['youth_sex'] = row[22].value

            data['youth_id_number'] = 'na'
            if row[17].value:
                data['youth_id_number'] = row[17].value  # UNHCR ID
            if row[18].value:
                data['youth_id_number'] = row[18].value  # Jordanian ID

            if row[19].value:
                data['youth_bayanati_ID'] = row[19].value  # Bayanati_ID
            else:
                data['youth_bayanati_ID'] = 'na'

            result = update_data(protocol=protocol, url=base_url, apifunc='/api/registration/{}/'.format(matched_registred_youth.id), token=token, data=data)

        except Exception as ex:
            print("---------------")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            pass


def import_assessments_as_registrations(filename, base_url, token, protocol='HTTPS'):

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet2']
    data = {}
    header = []
    index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            header.append((index, cell.value))
            index += 1

    for row in ws.rows:
        try:
            if row[0].value == 'start':
                continue
            data = {}

            #  pre assessment
            # if row[17].value == 'Yes':
            #     continue

            #  post assessment
            if row[17].value == 'Yes' or row[18].value == 'Yes':
                continue

            data['governorate'] = int(row[3].value) if row[3].value else None
            data['partner_organization'] = int(row[4].value) if row[4].value else None

            data['youth_first_name'] = row[6].value
            data['youth_last_name'] = row[7].value
            data['youth_father_name'] = row[8].value
            data['youth_sex'] = row[9].value

            data['youth_nationality'] = 1

            data['youth_birthday_year'] = ''
            data['youth_birthday_month'] = ''
            data['youth_birthday_day'] = ''
            if row[14].value:
                birthday = row[14].value.split('-')
                data['youth_birthday_year'] = int(birthday[0])
                data['youth_birthday_month'] = int(birthday[1])
                data['youth_birthday_day'] = int(birthday[2])

            result = post_data(protocol=protocol, url=base_url, apifunc='/api/registration/', token=token, data=data)
            registry = json.loads(result)

            #  pre
            # submit_assessment(header, row, 2, registry, base_url, token, protocol)

            #  post
            submit_assessment(header, row, 3, registry, base_url, token, protocol)

        except Exception as ex:
            print("---------------")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            pass


def submit_assessment(header, row, assessment_id, registry, base_url, token, protocol='HTTPS'):
    try:
        assessment_data = {}
        for key, value in header:
            assessment_data[value] = row[key].value

        assessment = {}
        assessment['registration'] = registry['id']
        assessment['youth'] = registry['youth_id']
        assessment['assessment'] = assessment_id
        assessment['data'] = assessment_data
        post_data(protocol=protocol, url=base_url, apifunc='/api/assessment-submission/', token=token, data=assessment)

    except Exception as ex:
        print("---------------")
        print("error: ", ex.message)
        print(json.dumps(assessment, cls=DjangoJSONEncoder))
        print("---------------")
        pass


@app.task
def import_assessments(assessment, filename, base_url, token, protocol='HTTPS'):

    from referral_platform.registrations.models import Registration

    wb = load_workbook(filename=filename, read_only=True)
    ws = wb['Sheet2']
    data = {}
    header = []
    index = 0
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            header.append((index, cell.value))
            index += 1

    for row in ws.rows:
        try:
            if row[0].value == 'start':
                continue
            registry = {}
            data = {}

            #  matched with registration
            # if not row[17].value == 'Yes':
            #     continue

            #  matched with pre
            if not row[18].value == 'No':
                continue

            birthday_day = ''
            birthday_month = ''
            birthday_year = ''
            if row[14].value:
                birthday = row[14].value.split('-')
                birthday_year = int(birthday[0])
                birthday_month = int(birthday[1])
                birthday_day = int(birthday[2])

            # id_number = generate_id(
            #     row[6].value,
            #     row[8].value,
            #     row[7].value,
            #     row[9].value,
            #     birthday_day,
            #     birthday_month,
            #     birthday_year,
            #     ""
            # )

            matched_registred_youth = Registration.objects.filter(
                youth__first_name=row[6].value,
                youth__father_name=row[8].value,
                youth__last_name=row[7].value,
                youth__sex=row[9].value,
                youth__birthday_year=birthday_year
            ).first()
            registry['id'] = matched_registred_youth.id
            registry['youth_id'] = matched_registred_youth.youth_id

            submit_assessment(header, row, assessment, registry, base_url, token, protocol)

        except Exception as ex:
            print("---------------")
            print("error: ", ex.message)
            print(json.dumps(data, cls=DjangoJSONEncoder))
            print("---------------")
            pass


class MyEncoder(json.JSONEncoder):

    def default(self, obj):
        if isinstance(obj, datetime.datetime) or isinstance(obj, datetime.date):
            return int(mktime(obj.timetuple()))

        return json.JSONEncoder.default(self, obj)

    def decode(self, obj):
        if isinstance(obj, datetime.datetime):
            return int(mktime(obj.timetuple()))

        return json.JSONEncoder.default(self, obj)


def post_data(protocol, url, apifunc, token, data):

    params = json.dumps(data, cls=MyEncoder)

    headers = {"Content-type": "application/json", "Authorization": token, "HTTP_REFERER": url, "Cookie": "token="+token}

    if protocol == 'HTTPS':
        conn = httplib.HTTPSConnection(url)
    else:
        conn = httplib.HTTPConnection(url)
    conn.request('POST', apifunc, params, headers)
    response = conn.getresponse()
    result = response.read()

    if not response.status == 201:
        if response.status == 400:
            raise Exception(str(response.status) + response.reason + response.read())
        else:
            raise Exception(str(response.status) + response.reason)

    conn.close()

    return result


def update_data(protocol, url, apifunc, token, data):

    params = json.dumps(data, cls=MyEncoder)

    headers = {"Content-type": "application/json", "Authorization": token, "HTTP_REFERER": url, "Cookie": "token="+token}

    if protocol == 'HTTPS':
        conn = httplib.HTTPSConnection(url)
    else:
        conn = httplib.HTTPConnection(url)
    conn.request('PATCH', apifunc, params, headers)
    response = conn.getresponse()
    result = response.read()

    if not response.status == 200:
        if response.status == 400:
            raise Exception(str(response.status) + response.reason + response.read())
        else:
            raise Exception(str(response.status) + response.reason)

    conn.close()

    return result
